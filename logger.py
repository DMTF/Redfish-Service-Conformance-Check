# Copyright Notice:
# Copyright 2016-2017 Distributed Management Task Force, Inc. All rights reserved.
# License: BSD 3-Clause License. For full text see link: https://github.com/DMTF/Redfish-Service-Conformance-Check/blob/master/LICENSE.md

###################################################################################################
# Name: logger
# Description: This module contains Log class for logging purpose of this tool based on each inst-
#   -ance of SUT. It contains limited functionality and functionalities could be expanded
#   What works: Functions related to manipulating assertion excel sheet provided with the tool
#   (rf-assertions-run.xlsx in assertions folder) updating the assertion status against each
#   assertion id plus some additional comments into the assertion excel sheet and places it in a
#   result folder. It also produces a detailed log text file for the assertions.
#
# Verified/operational Python revisions (Windows OS) :
#       2.7.10
#       3.4.3
#
# Initial code released : 01/2016
#   Steve Krig      ~ Intel
#   Fatima Saleem   ~ Intel
#   Priyanka Kumari ~ Texas Tech University
###################################################################################################

import argparse
import base64
import warnings
import shutil
from datetime import datetime
import sys
import os

## openpyxl is not a default install for python - you will need to install it using 'pip'... 
# -- to install it...
# cd to where your python.exe is and into the /Scripts subdirectory to run pip.exe... and if 
# you are behind a firewall you may need to specify a proxy server to get to the installation 
# server..
# dos_box>  pip --proxy <hostname>:<port> install openpyxl

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Side, Border, colors, PatternFill, Font

###################################################################################################
# Class: Log
#   This class contatins text and excel sheet logging related functions below are controls for 
#   accessing the spreadsheet cells...
###################################################################################################     
class Log:         
    def __init__(self):
        # tracking tool release revision with a date stamp -  month:day:year   
        self.RedfishServiceCheck_Revision = "07.11.16"
        # assertion status
        self.PASS = 'PASS'
        self.WARN = 'WARN'
        self.FAIL = 'FAIL'
        self.INCOMPLETE = 'PASS (Incomplete). Check Log for details' # to remove

        # Redfish latest spec url
        self.RedfishSpecHyperlinkPath = 'https://www.dmtf.org/sites/default/files/standards/documents/DSP0266_1.0.2.pdf'
        # get current script file directory, log files/folders should be in this directory
        self.ScriptDirectory = os.path.dirname(__file__)
        # Folder name where log files reside
        self.LogDestFolder = 'logs'
        # Path of excel sheet destination folder for logging purpose
        self.LogDestinationPath = os.path.join(self.ScriptDirectory, self.LogDestFolder)
        # Following will get set based on the SUT DisplayName from the properties.json when the tool runs...
        # Path of SUTs destination log folder within the general logs folder
        self.SUT_log_Folder = None

    ###############################################################################################
    # Name: init_xl
    #   initializes variables required for assertion excel sheet control functions and necc paths 
    #   for source and destination excel sheet
    ###############################################################################################
    def init_xl(self):  
        self.XlAssertionSheet = 0
        self.XlAssertionWb = 0

        self.RedfishHyperlinkRow = 4
        self.RedfishHyperlinkCol = 2
        
        ## assertions xlxs spreadsheet support - color fills for pass/warn/fail status
        self.xl_PASS = PatternFill(fill_type='solid', start_color=colors.GREEN, end_color=colors.GREEN)
        self.xl_WARN = PatternFill(fill_type='solid', start_color=colors.YELLOW, end_color=colors.YELLOW)
        self.xl_FAIL = PatternFill(fill_type='solid', start_color=colors.RED, end_color=colors.RED)
        self.xl_INCOMPLETE = PatternFill(fill_type='solid', start_color=colors.GREEN, end_color=colors.GREEN) # to remove or fix
        
        ## alignment and word wrap - used when writing the log header to the top of the spreadsheet
        self.xl_Alignment = Alignment(horizontal='center',\
				    vertical='center',\
				    text_rotation=0,\
				    wrap_text=True,\
				    shrink_to_fit=False,\
				    indent=0)        

        ## log header row/col location
        self.assertion_logHeaderRow = 1
        self.assertion_logHeaderCol = 2

        ## assertions xlxs row/col assignments
        self.Assertion_ID_column = 1
        self.Assertion_description_column = 2
        self.Assertion_comment_column = 5

        ## status counters
        self.Assertion_Counter = {\
            # status counters for number of assertions run
            self.PASS : 0,\
            self.WARN : 0,\
            self.FAIL : 0,\
            self.INCOMPLETE : 0
        }

        ## set to ID of Assertion currently being run - this becomes the 'key' into the spreadsheet 
        ## to locate the assertion text
        self.AssertionID = '0.0.0'

        # Following are the expected folder/file names by the tool. If these folder/file names are 
        # changed outside the tool, Make sure to update the following..
        # Folder name where original assertion excel sheet reside
        self.AssertionSrcFolder = 'assertions'
        # Name of master assertions excel sheet
        self.xl_RunFileName = 'rf-assertions-run.xlsx' 
        # Name of text log file
        self.TxtFileName = 'rf-assertions-log'
        # Following will get set based on the SUT DisplayName from the properties.json when the tool runs...
        # Path of SUTs destination log file within the SUTs dest log folder
        self.SUT_XlDestPath = None

        # Path of excel sheet source folder
        self.AssertionSrcPath = os.path.join(self.ScriptDirectory, self.AssertionSrcFolder)

        # checking if logging destination folder exists at the path, if not tool tries to create one..
        if not os.path.isdir(self.LogDestinationPath):
            try:
                os.makedirs(self.LogDestinationPath)
                os.chmod(self.LogDestinationPath, 0o777)
            except Exception as inst:
                    print('Operational ERROR - Tool was unable to create a Log folder. Create a folder named: %s in the script directory: %s manually and try running the tool again.' % (self.LogDestFolder, self.ScriptDirectory))
                    print (type(inst))     # the exception instance
                    print (inst.args)
                    exit(0)

        # Path of the copy of the assertion master excel sheet which gets marked during the test run
        self.XlRunPath = os.path.join(self.AssertionSrcPath, self.xl_RunFileName)
          
        # text log file for assertions logging 
        self.TextLogPath = None
        self.TextLogHandle = 0

    ###############################################################################################
    # Name: init_logfile(log_name)
    #   Takes a file name and initialies a new Text log file. Optionally to take SUTs property
    #   Returns: 
    #       If successfully created, file handle to access file and final path of the log file.
    #       Else None
    ###############################################################################################
    def init_logfile(self, log_name, SUT_prop = None):
        log_folder = None
        # checking if logging destination folder exists at the path, if not tool tries to create one..
        if not os.path.isdir(self.LogDestinationPath):
            try:
                os.makedirs(self.LogDestinationPath)
                os.chmod(self.LogDestinationPath, 0o777)
            except Exception as inst:
                    print('Operational ERROR - Tool was unable to create a Log folder. Create a folder named: %s in the script directory: %s manually and try running the tool again.' % (self.LogDestFolder, self.ScriptDirectory))
                    print (type(inst))     # the exception instance
                    print (inst.args)
                    exit(0)
            else:
                if not self.SUT_log_Folder and SUT_prop:
                    log_Folder = os.path.join(self.LogDestinationPath, SUT_prop['DisplayName'])
                    if self.SUT_log_Folder:
                        if not os.path.isdir(self.SUT_log_Folder):
                            try:
                                os.makedirs(self.SUT_log_Folder)
                                os.chmod(self.SUT_log_Folder,0o777)
                            except Exception as inst:
                                print('Operational ERROR - Tool was unable to create a Log folder for current SUT. Try placing a folder named: %s in the script directory: %s manually and try running the tool again.' % (SUT_prop['DisplayName'], self.LogDestinationPath))
                                print (type(inst))     # the exception instance
                                print (inst.args)
                                exit(0)

                    if not SUT_prop:
                        log_Folder = self.LogDestinationPath

                if log_folder:
                    ## create a unique log file name 
                    dstr = str(datetime.now().strftime("%Y%m%d-%H%M%S"))               
                    log_file = dstr + '_' + log_name
                    log_path = os.path.join(log_Folder, log_file)
                    try:
                        # create a new log file to start logging 
                        log_handle = open(log_path, 'a')
                    except:
                        print('Operational ERROR - Tool was unable to create a Log file with name: %s at destination: %s' % (log_name, log_Folder))
                        print (type(inst))     # the exception instance
                        print (inst.args)
                    else:
                        return log_handle, log_path

        print('Operational ERROR - Tool was unable to create a Log file with name: %s' % (log_name, log_Folder))
        return None, None                        
                
    ###############################################################################################
    # Name: open_assertions_xl()                          
    #   open the xlxs file containing the redifish assertions at the specified pathname
    # Returns:
    #   If no error, a handle which can be used to access the assertions spreadsheet for reads/writes
    #   else 0
    ###############################################################################################
    def open_assertions_xl(self) : 
        # load the assertion list xls file - python throws some cryptic warning here..
        #  so code in place to ignore  the warning...
        warnings.simplefilter("ignore")
        try:   
            self.XlAssertionWb = load_workbook(filename=self.SUT_XlDestPath)

        except:
            return 0

        # get a 'handle to the assertions sheet
        self.XlAssertionSheet = self.XlAssertionWb.get_active_sheet()

        return 1
    #
    ## end open_assertions_xl()

    ##
    # save changes to the assertion file
    ##
    def save_assertions_xl(self):
        try:
            self.XlAssertionWb.save(self.SUT_XlDestPath)
            ## success
            return 1
        except:
            # unable to update the spreadsheet -- user probably has it open
            return 0

    ###############################################################################################
    # Name: assertion_id_row(assertion_id)                        
    #   Takes an assertion id as a key (string), locate the row in the spreadsheet containing
    #   the assertion and return the row number
    # Returns:
    #   on assertion id match...  return the row number of the assertion in the spreadsheet/xl file; 
    #   else 0                
    ###############################################################################################
    def assertion_id_row(self, assertion_id):

        asx_handle = self.XlAssertionSheet

        #find the assertion id in the xls file...
        row_cnt=1
        for row in asx_handle.iter_rows(row_offset=1):
            row_assertion_id = asx_handle.cell(row=row_cnt, column=self.Assertion_ID_column).value
            if row_assertion_id == assertion_id :
                ## success
                return row_cnt    
            row_cnt += 1

        ## failure        
        print('Operational ERROR unable to find Assertion ID %s in the assertion xlxs file' % assertion_id)

        return 0        
    #
    ## end assertion_id_row()

    ###############################################################################################
    # Name: assert_xl(assertion_id, pwf_stat)                        
    #   Takes an assertion id as a key (string), locate the row in the spreadsheet containing
    #   the assertion description text and return that text - mark pass fail warn status in the 
    #   spreadsheet depending upon setting of pwf_stat to 'PASS', 'FAIL', 'WARN' or None;
    #       - if 'NONE' then just return the assertion description from the spreadsheet for the 
    #         assertion ID match  
    #       - if 'PASS' 'FAIL' or 'WARN' then mark an assertion cell in the assertion xl file 
    #         green/yellow/red depending on pass/warn/fail
    # Returns:
    #   on assertion id match...  return the description of the assertion as read from
    #       the spreadsheet; else return ' '               
    ###############################################################################################
    def assert_xl(self, assertion_id, pwf_stat):

        asx_handle = self.XlAssertionSheet

        assert_descr = ' '
        
        #find a particular assertion in the xls file...
        zrow = self.assertion_id_row(assertion_id)
        if (zrow > 0):
            assert_descr = asx_handle.cell(row=zrow, column=self.Assertion_description_column).value

            # mark the assertion id cell with pass/warn/fail status...
            if (pwf_stat == self.PASS):
                asx_handle.cell(row=zrow, column=self.Assertion_ID_column).fill = self.xl_PASS
            elif (pwf_stat == self.WARN):
                asx_handle.cell(row=zrow, column=self.Assertion_ID_column).fill = self.xl_WARN
            elif (pwf_stat == self.FAIL):
                asx_handle.cell(row=zrow, column=self.Assertion_ID_column).fill = self.xl_FAIL
            elif (pwf_stat == self.INCOMPLETE):
                asx_handle.cell(row=zrow, column=self.Assertion_ID_column).fill = self.xl_PASS

            self.save_assertions_xl()
     
        return assert_descr        
    #
    ## end assert_xl()

    ###############################################################################################
    # Name: assertion_log(log_control, log_string, SUT_prop = None, service_root = None)                                            
    #   Takes Log control key (OPEN, CLOSE, XL_COMMENT, TX_COMMENT and line) and log message string
    #   and based on the assertionid set thru assertion, it updates the log. It also updates the 
    #   color for the ID cell in the excel sheet. Green for pass, Yellow for warninf and red for 
    #   fail
    #
    #   - OPEN: REQUIRED params: Properties of SUT and service root url.
    #           Makes a copy the excel sheet for the SUT w/current time and updates headers in the
    #           spreadsheet. Also creates/opens a text log file for the SUT.
    #   - CLOSE: Makes appropriate updates of total assertions status in files and closes it
    #   - XL_COMMENT: Updates status and string in excel sheet against the given assertion id's row
    #   - TX_COMMENT: Appends status and string in text file w/assertion id, if provided
    #   - line: Prints status and string on command line 
    #               
    # Return: 0 on failure; 1 on success
    ################################################################################################
    def assertion_log(self, log_control, log_string, SUT_prop = None, service_root = None) :       
        assertion_id = self.AssertionID
        ##
        # handle open/close of the log files
        #
        if (log_control == 'OPEN' and SUT_prop and service_root):
            ## open the assertion log files and write a unique test header for this run
            self.SUT_log_Folder = os.path.join(self.LogDestinationPath, SUT_prop['DisplayName'])
            if not os.path.isdir(self.SUT_log_Folder):
                try:
                    os.makedirs(self.SUT_log_Folder)
                    os.chmod(self.SUT_log_Folder,0o777)
                except Exception as inst:
                    print('Operational ERROR - Tool was unable to create a Log folder for current SUT. Try placing a folder named: %s in the script directory: %s \
                    manually and try running the tool again.' % (SUT_prop['DisplayName'], self.LogDestinationPath))
                    print (type(inst))     # the exception instance
                    print (inst.args)
                    exit(0)

            # open the text log file
            self.TextLogPath = os.path.join(self.SUT_log_Folder, self.TxtFileName)
            try:
                self.TextLogHandle = open(self.TextLogPath, 'a')
            except Exception as inst:
                print('Operational ERROR - unable to create/open the text log file %s. Try placing a text file ' % self.TextLogPath)
                print (type(inst))     # the exception instance
                print (inst.args)
                exit(0)

            #datetime string
            dstr = str(datetime.now().strftime("%Y%m%d-%H%M%S"))
           
            ## create a copy of the master assertion xlxs file for this SUT and open it
            #self.SUT_XlDestPath = os.path.join(self.SUT_log_Folder , self.xl_RunFileName)
            self.SUT_XlDestPath = os.path.join(self.SUT_log_Folder , dstr + '_' + self.xl_RunFileName)
            try:
                shutil.copyfile(self.XlRunPath, self.SUT_XlDestPath)
            except Exception as inst:
                print('Operational ERROR unable to create %s\n - make sure you have %s in your local directory and\n  %s is not already open' % (self.SUT_XlDestPath, self.XlRunPath, self.XlRunPath))
                print (type(inst))     # the exception instance
                print (inst.args)
                exit(0)

            self.open_assertions_xl()

            if (self.XlAssertionSheet == 0):
                print('Operational ERROR - unable to open the assertions %s' % self.XlRunPath)
                return(0)

            ## create a unique log header 
            log_header_src = ('Redfish Service Check Tool Revision: %s : ' % self.RedfishServiceCheck_Revision)\
                + dstr + ' : '\
                + SUT_prop['DisplayName'] \
                + ':'+ SUT_prop['DnsName'] \
                + service_root

            # write log header to the log files... and console
            self.TextLogHandle.write('\nASSERTION RUN--->' + log_header_src + '<---' + '\n')
            print('\n' + log_header_src + '\n')

            # log header to the xls file
            self.XlAssertionSheet.cell(row=self.assertion_logHeaderRow, column=self.assertion_logHeaderCol).fill = self.xl_PASS
            self.XlAssertionSheet.cell(row=self.assertion_logHeaderRow, column=self.assertion_logHeaderCol).value = log_header_src
            self.XlAssertionSheet.cell(row=self.assertion_logHeaderRow, column=self.assertion_logHeaderCol).alignment = self.xl_Alignment

            # hyperlink to the Redfish spec
            self.XlAssertionSheet.cell(row=self.RedfishHyperlinkRow, column=self.RedfishHyperlinkCol).hyperlink = self.RedfishSpecHyperlinkPath            

            self.save_assertions_xl()

            # initialize the assertion counters
            self.Assertion_Counter[self.PASS] = 0
            self.Assertion_Counter[self.FAIL] = 0
            self.Assertion_Counter[self.WARN] = 0
            self.Assertion_Counter[self.INCOMPLETE] = 0

            # 
            ## End open/initialize log files

        elif (log_control == 'CLOSE'):
            self.AssertionID = None
            # log the tally of pass/warn/fail stats and close the log files
            completion_str = '\n Assertions Stats:\n Passed= %s Warn= %s Failed= %s \n Total Assertions Run= %s' % (str(self.Assertion_Counter[self.PASS] + self.Assertion_Counter[self.INCOMPLETE]), str(self.Assertion_Counter[self.WARN]), str(self.Assertion_Counter[self.FAIL]), str(self.Assertion_Counter[self.PASS] + self.Assertion_Counter[self.INCOMPLETE] + self.Assertion_Counter[self.WARN] + self.Assertion_Counter[self.FAIL]))

            self.assertion_log('line', completion_str)
            self.assertion_log('XL_LOG_HEADER', completion_str)

            self.TextLogHandle.close()

            print(' Assertions check successfully completed. Please see assertion spreadsheet: %s for checked assertions summary and log files: %s and %s for detailed log\n' % (self.XlRunPath, self.SUT_XlDestPath, self.TextLogPath))
        #
        # end of handling open/close of log files
        ##

        # log an assetion id tag at the start of an assertion
        elif (log_control == 'BEGIN_ASSERTION'):
            assert_string = '\n---> Assertion: ' + assertion_id + '\n'
            self.TextLogHandle.write(assert_string)
            print(assert_string)

        # write a string to the header column of the assertion spreadsheet 
        elif (log_control == 'XL_LOG_HEADER'):
            # add a line to the header in the xls file
            log_string = self.XlAssertionSheet.cell(row=self.assertion_logHeaderRow, column=self.assertion_logHeaderCol).value + ' :: ' + log_string
            self.XlAssertionSheet.cell(row=self.assertion_logHeaderRow, column=self.assertion_logHeaderCol).value = log_string

            self.save_assertions_xl()
        
        # pass fail to the text log file and color code the assertion row in the assertion spreadsheet
        # and increment pass/warn/fail counters
        elif (log_control == self.PASS) or (log_control == self.WARN) or (log_control == self.FAIL) or (log_control == self.INCOMPLETE):
            # mark/color the assertion id column of the spreadsheet and get the description text
            # for the assertion 
            assertion_description = self.assert_xl(assertion_id, log_control)

            # log pass/fail status to the text log 
            if (log_control != self.PASS or log_control != self.INCOMPLETE):
                # include the assertion description in the text log
                log_string =  ('Assertion Description: %s\n<--- Assertion %s: %s\n' % (assertion_description.encode('utf-8'), self.AssertionID, log_control))
            else:
                log_string =  ('<--- Assertion %s: %s\n' % (self.AssertionID, log_control))

            self.TextLogHandle.write(log_string)
            print(log_string)
                
            # increment the pass/warn/fail counter
            self.Assertion_Counter[log_control] += 1

        # write string to the comment row/column of the assertion spreadsheet for this assertion id
        if ((log_control == 'XL_COMMENT') or (log_control == 'line')):
            if self.AssertionID:
                xl_row = self.assertion_id_row(assertion_id)
                if (xl_row > 0):
                    current_cell_value = self.XlAssertionSheet.cell(row=xl_row, column=self.Assertion_comment_column).value
                    if current_cell_value == None:
                        current_cell_value = ''
                    self.XlAssertionSheet.cell(row=xl_row, column=self.Assertion_comment_column).value = current_cell_value + log_string
                self.save_assertions_xl()

        # write a line into the text log file and/or console
        if ((log_control == 'line') or (log_control == 'TX_COMMENT')) :
            if log_string == None:
                log_string = ""
            
            #  output to the text log file
            self.TextLogHandle.write(log_string + '\n')

            # output to the console
            if (log_control != 'TX_COMMENT'):
                print(log_string +'\n')

        # success
        return(1) 
    #
    ## end _assertion_log

    ###############################################################################################
    # Name: schema_log()         WIP                                   
    #   WIP, based on a log file created thru init_logfile, we can access it 
    #   with the handle it returns and use control options OPEN, CLOSE, COMMENT to update it with 
    #   any logging info
    ###############################################################################################
    def schema_log(self, log_control, log_handle, log_path, sut_prop = None) :
        ##
        # handle open/close of the log files
        #
        if (log_control == 'OPEN'):
            ## create a unique log header 
            dstr = str(datetime.now().strftime("%Y%m%d-%H%M%S"))
            log_header_src = ('Redfish Service Check Tool Revision: %s : ' % self.RedfishServiceCheck_Revision)\
                + dstr + ' : '
            log_handle.write(log_header_src + '\n')
            # 
            ## End open/initialize log files
            log_handle.close()
        #
        # end of handling open/close of log files
        #
        # write a line into the text log file and/or console
        if ((log_control == 'line') or (log_control == 'COMMENT')) :
            if log_string == None:
                log_string = ""
            
            #  output to the text log file
            log_handle.write(log_string + '\n')

            # output to the console
            if (log_control != 'COMMENT'):
                print(log_string +'\n')

        # success
        return(1) 
    #
    ## end schema_log

    ###############################################################################################
    # Name: status_fixup()                                            
    #   Takes 2 status strings and returns status giving precendence in the order FAIL, WARN, INCOMPLETE, PASS
    ###############################################################################################
    def status_fixup(self, assertion_status, assertion_status_):
        if assertion_status == self.FAIL and (assertion_status_ == self.WARN or assertion_status_ == self.INCOMPLETE or assertion_status_ == self.PASS):
            return assertion_status
        if assertion_status == self.WARN and (assertion_status_ == self.INCOMPLETE or assertion_status_ == self.PASS):
            return assertion_status
        if assertion_status == self.INCOMPLETE and (assertion_status_ == self.PASS):
            return assertion_status
        
        return assertion_status_
