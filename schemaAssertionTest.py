import urllib3
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from time import gmtime, strftime

#P = input('Enter port number: ')

http = urllib3.PoolManager()
wb = load_workbook('SchemaAssertions.xlsx')
ws = wb.active

def setBG(s, e, c):
    for row in ws[s:e]:
        for cell in row:
            cell.fill = PatternFill(start_color= c, end_color= c, fill_type = 'solid')
            

def ACCO100(): 
    
    print('Testing Rule# ACCO100')
    r = http.request('GET', 'http://127.0.0.1:8001/redfish/v1/AccountService')
    
    if r.status == 200:
        setBG('A3', 'C3', '71fc62')
        ws['C3'] = 'Assertion passed.'

    else:
        setBG('A3', 'C3', 'ff4d4d')
        ws['C3'] = 'Assertion failed due to the absence of the resource.'

   


def ACCO101():
    
    print('Testing Rule# ACCO101')
    r = http.request('GET', 'http://127.0.0.1:8001/redfish/v1/AccountService')

    if r.status == 200:
        
        data =  json.loads(r.data.decode('utf-8')) 
         
        if data.get('ServiceEnabled') == True or data.get('ServiceEnabled') == False :
            setBG('A4', 'C4', '71fc62')
            ws['C4'] = 'Assertion passed.'
        
        else:
            setBG('A4', 'C4', 'ff4d4d')
            ws['C4'] = 'Assertion failed due to the absence of the property or non boolean property value type.'
    
    else:
        setBG('A4', 'C4', 'ff4d4d')
        ws['C4'] = 'Assertion failed due to the absence of the resource.'



def ACCO102():
    
    print('Testing Rule# ACCO102')
    r = http.request('GET', 'http://127.0.0.1:8001/redfish/v1/AccountService')

    if r.status == 200:
        data =  json.loads(r.data.decode('utf-8'))  
        
        if r.status == 200 and isinstance(data.get('AuthFailureLoggingThreshold'), int):
            setBG('A5', 'C5', '71fc62')
            ws['C5'] = 'Assertion passed.'
        
        else:
            setBG('A5', 'C5', 'ff4d4d')
            ws['C5'] = 'Assertion failed due to the absence of the property or non number property value type.'
    
    else:
        setBG('A5', 'C5', 'ff4d4d')
        ws['C5'] = 'Assertion failed due to the absence of the resource.'    
    
#Testing
    
ACCO100()

ACCO101()

ACCO102()

tm = strftime(" %H_%M_%S", gmtime()) 


tmst = 'SchemaAssertionsLog' + tm + '.xlsx'

wb.save(tmst)


    

